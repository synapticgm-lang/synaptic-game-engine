from pathlib import Path
import json
from openpyxl import load_workbook

ROOT = Path('/home/ubuntu/SynapticGM_own_ai_cashflow_gapfill_2026-08-18')
source = ROOT / 'deliverables' / 'SynapticGM_own_ai_cashflow_gapfill_2026-08-18_G2_G3_G4_cashflow_calculator.xlsx'
recalc = ROOT / 'model' / 'recalculated' / source.name

wb_formula = load_workbook(source, data_only=False)
wb_value = load_workbook(recalc, data_only=True)

result = {
    'expected_sheets': ['Read Me', 'Assumptions', 'Cashflow', 'Sensitivity', 'Image Ladder', 'Sources'],
    'actual_sheets': wb_formula.sheetnames,
    'sheet_order_ok': wb_formula.sheetnames == ['Read Me', 'Assumptions', 'Cashflow', 'Sensitivity', 'Image Ladder', 'Sources'],
    'cashflow_sample_formulas': {},
    'cashflow_recalculated_values': {},
    'formula_error_cells': [],
    'assumption_inputs_with_comments': 0,
    'assumption_inputs_without_comments': [],
}

wsf = wb_formula['Cashflow']
wsv = wb_value['Cashflow']
for coord in ['D9', 'E9', 'F9', 'Q9', 'R9', 'D14', 'H21', 'Q21', 'R21']:
    result['cashflow_sample_formulas'][coord] = wsf[coord].value
    result['cashflow_recalculated_values'][coord] = wsv[coord].value

for ws in wb_value.worksheets:
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith('#'):
                result['formula_error_cells'].append(f'{ws.title}!{cell.coordinate}={cell.value}')

wsa = wb_formula['Assumptions']
for row in wsa.iter_rows(min_row=8, max_row=wsa.max_row, min_col=4, max_col=4):
    cell = row[0]
    if isinstance(cell.value, (int, float)):
        if cell.font.color and cell.font.color.type == 'rgb' and cell.font.color.rgb and cell.font.color.rgb.endswith('0000FF'):
            if cell.comment:
                result['assumption_inputs_with_comments'] += 1
            else:
                result['assumption_inputs_without_comments'].append(cell.coordinate)

print(json.dumps(result, indent=2, default=str))
