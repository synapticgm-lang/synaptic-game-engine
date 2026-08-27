from pathlib import Path
from math import ceil
import csv
import json
import os

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.comments import Comment
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
import matplotlib.pyplot as plt

ROOT = Path('/home/ubuntu/SynapticGM_own_ai_cashflow_gapfill_2026-08-18')
OUT_DIR = ROOT / 'deliverables'
MODEL_DIR = ROOT / 'model'
OUT_DIR.mkdir(parents=True, exist_ok=True)

PREFIX = 'SynapticGM_own_ai_cashflow_gapfill_2026-08-18'
XLSX_PATH = OUT_DIR / f'{PREFIX}_G2_G3_G4_cashflow_calculator.xlsx'
CSV_PATH = MODEL_DIR / f'{PREFIX}_cashflow_results.csv'
JSON_PATH = MODEL_DIR / f'{PREFIX}_cashflow_results.json'
PNG_PATH = OUT_DIR / f'{PREFIX}_cost_comparison_chart.png'

# All assumptions are planning inputs. Vendor rates are fresh list prices observed 18 Aug 2026.
A = {
    'usd_per_gbp': 1.3553,
    'gbp_per_usd': 1 / 1.3553,
    'days_per_month': 30,
    'hours_per_month': 730,
    'weeks_per_month': 52 / 12,
    'mau_list': [100, 1000, 10000],
    'turns_per_user_month': 40,
    'input_tokens_turn': 5000,
    'output_tokens_turn': 500,
    'cache_hit_rate': 0.30,
    'cache_write_share_when_enabled': 0.05,
    'retry_rate_api': 0.05,
    'retry_rate_hosted_gate': 0.03,
    'api_fallback_rate_hosted_gate': 0.01,
    'api_fallback_rate_warden': 0.01,
    'api_fallback_rate_full_7b': 0.85,
    'api_fallback_rate_full_70b': 0.10,
    'peak_factor': 10,
    'warden_gpu_seconds_turn': 0.60,
    'warden_gpu_usd_hour': 0.27,
    'warden_serverless_usd_hour': 0.69,
    'full_7b_tps': 160,
    'full_7b_gpu_usd_hour': 0.27,
    'full_70b_tps': 50,
    'full_70b_gpu_usd_hour': 1.39,
    'operator_gbp_hour': 40,
    'image_usd_per_1mp': 0.03,
    'weekly_active_share': 0.60,
    'weekly_image_cap_per_wau': 1.0,
    'art_skip_rate': 0.70,
    'image_delivery_success': 0.90,
    'memorable_keep_rate': 0.65,
    'gate_input_tokens_turn': 2200,
    'gate_output_tokens_turn': 80,
}

# Representative direct narrator: Anthropic Sonnet 5, current direct API table.
RATES = {
    'narrator_input': 2.00,
    'narrator_cache_read': 0.20,
    'narrator_cache_write': 2.50,
    'narrator_output': 10.00,
    'openrouter_loading': 0.055,
    'gate_input': 0.05,
    'gate_cache_read': 0.01,
    'gate_output': 0.20,
}

OPS = {
    'A_openrouter_only': {'hours': 1, 'eval_gbp': 25},
    'B_direct_cache': {'hours': 2, 'eval_gbp': 40},
    'C_hosted_gates_mid': {'hours': 3, 'eval_gbp': 60},
    'D_selfhost_warden': {'hours': 6, 'eval_gbp': 100},
    'E1_selfhost_7b': {'hours': 8, 'eval_gbp': 150},
    'E2_selfhost_70b': {'hours': 14, 'eval_gbp': 300},
}

SOURCE = {
    'usd_per_gbp': 'Source: Bank of England daily spot exchange rate; 17 Aug 2026; accessed 18 Aug 2026. £1 = $1.3553. Planning translation only. https://www.bankofengland.co.uk/boeapps/database/Rates.asp',
    'narrator': 'Source: Anthropic Pricing; accessed 18 Aug 2026. Claude Sonnet 5: $2/MTok input, $0.20 cache hit, $2.50 5m cache write, $10/MTok output. https://docs.anthropic.com/en/docs/about-claude/pricing',
    'openrouter': 'Source: OpenRouter Pricing; accessed 18 Aug 2026. Pay-as-you-go platform fee displayed as 5.5%. Planning gross-up, verify account billing before commitment. https://openrouter.ai/pricing',
    'gate': 'Source: Fireworks Serverless Pricing; accessed 18 Aug 2026. NVIDIA Nemotron 3.5 Lightning 30B A3B: $0.05 input / $0.01 cached input / $0.20 output per MTok. https://docs.fireworks.ai/serverless/pricing',
    'runpod_24': 'Source: RunPod Pricing; accessed 18 Aug 2026. Community Cloud RTX A5000 24GB $0.27/GPU-hour; 24GB Serverless L4/A5000/3090/MIG $0.69/GPU-hour. https://www.runpod.io/pricing',
    'runpod_80': 'Source: RunPod Pricing; accessed 18 Aug 2026. A100 PCIe 80GB Community Cloud $1.39/GPU-hour. https://www.runpod.io/pricing',
    'image': 'Source: BFL Pricing and OpenRouter FLUX.2 Pro page; accessed 18 Aug 2026. FLUX.2 Pro first output MP $0.03; reference input $0.015/MP. https://docs.bfl.ai/quick_start/pricing ; https://openrouter.ai/black-forest-labs/flux.2-pro',
    'planning': 'Source: SynapticGM planning assumption; 18 Aug 2026. Editable model input; not a vendor price or forecast.',
}

SCENARIOS = [
    ('A_openrouter_only', 'A — OpenRouter-only', 'OpenRouter routes a mid narrator; no custom gate or GPU.'),
    ('B_direct_cache', 'B — Direct APIs + prompt cache', 'Direct mid narrator with cache-aware stable prefix; no custom gate or GPU.'),
    ('C_hosted_gates_mid', 'C — Hosted low-cost gates + mid narrator', 'Hosted low-cost pre/post gate plus direct mid narrator.'),
    ('D_selfhost_warden', 'D — Self-host Continuity Warden only', 'Direct mid narrator plus one warm 24GB Warden Pod; hosted-gate fallback.'),
    ('E1_selfhost_7b', 'E1 — Self-host full 7B narrator (production fallback)', 'A 7B self-hosted narrator with an explicit 85% paid-API quality/load fallback until it proves equivalence on the full suite.'),
    ('E2_selfhost_70b', 'E2 — Self-host full 70B narrator', 'A 70B service-aspirational narrator on A100 80GB capacity; API fallback for residual failures.'),
]


def cache_components(input_rate, cache_rate, write_rate, hit_rate=None):
    """Return effective cost per million input tokens with cache-hit and write shares."""
    h = A['cache_hit_rate'] if hit_rate is None else hit_rate
    w = A['cache_write_share_when_enabled'] if h > 0 else 0.0
    if h + w > 1:
        raise ValueError('cache hit plus write share must not exceed 1')
    return h * cache_rate + w * write_rate + (1 - h - w) * input_rate


def narrator_usd_per_turn(openrouter=False, hit_rate=None):
    multiplier = 1 + RATES['openrouter_loading'] if openrouter else 1.0
    e_in = cache_components(RATES['narrator_input'], RATES['narrator_cache_read'], RATES['narrator_cache_write'], hit_rate)
    return multiplier * ((A['input_tokens_turn'] / 1_000_000) * e_in + (A['output_tokens_turn'] / 1_000_000) * RATES['narrator_output'])


def gate_usd_per_turn(hit_rate=None):
    # Fireworks cache-write pricing is not public as a fixed multiplier; use no cache-write premium and conservatively model cache reads only.
    h = A['cache_hit_rate'] if hit_rate is None else hit_rate
    e_in = h * RATES['gate_cache_read'] + (1 - h) * RATES['gate_input']
    return ((A['gate_input_tokens_turn'] / 1_000_000) * e_in
            + (A['gate_output_tokens_turn'] / 1_000_000) * RATES['gate_output'])


def art_usd_month(mau, skip_rate=None):
    s = A['art_skip_rate'] if skip_rate is None else skip_rate
    attempts = mau * A['weekly_active_share'] * A['weekly_image_cap_per_wau'] * A['weeks_per_month'] * (1 - s)
    return attempts * A['image_usd_per_1mp']


def operator_cost_gbp(scenario_key):
    return OPS[scenario_key]['hours'] * A['operator_gbp_hour']


def capacity(mau, tps, min_gpu=1):
    turns = mau * A['turns_per_user_month']
    avg_turns_per_sec = turns / (A['days_per_month'] * 24 * 3600)
    peak_turns_per_sec = avg_turns_per_sec * A['peak_factor']
    return max(min_gpu, ceil(peak_turns_per_sec * A['output_tokens_turn'] / tps))


def warden_capacity(mau):
    turns = mau * A['turns_per_user_month']
    avg_turns_per_sec = turns / (A['days_per_month'] * 24 * 3600)
    return max(1, ceil(avg_turns_per_sec * A['peak_factor'] * A['warden_gpu_seconds_turn']))


def compute(scenario_key, mau, hit_rate=None, retry_rate=None, art_skip=None):
    turns = mau * A['turns_per_user_month']
    fx = A['gbp_per_usd']
    direct_turn = narrator_usd_per_turn(False, hit_rate)
    openrouter_turn = narrator_usd_per_turn(True, hit_rate)
    gate_turn = gate_usd_per_turn(hit_rate)
    art_gbp = art_usd_month(mau, art_skip) * fx
    out = {
        'scenario_key': scenario_key,
        'mau': mau,
        'turns_month': turns,
        'narrator_gbp': 0.0,
        'gate_gbp': 0.0,
        'gpu_gbp': 0.0,
        'operator_gbp': operator_cost_gbp(scenario_key),
        'eval_gbp': OPS[scenario_key]['eval_gbp'],
        'fallback_gbp': 0.0,
        'art_gbp': art_gbp,
        'gpu_count': 0,
        'gpu_active_hours': 0.0,
        'gpu_paid_hours': 0.0,
        'gpu_idle_hours': 0.0,
        'notes': '',
    }
    if scenario_key == 'A_openrouter_only':
        rr = A['retry_rate_api'] if retry_rate is None else retry_rate
        out['narrator_gbp'] = turns * openrouter_turn * (1 + rr) * fx
        out['notes'] = 'OpenRouter direct narrator cost includes 5.5% planning loading and retry rate.'
    elif scenario_key == 'B_direct_cache':
        rr = A['retry_rate_api'] if retry_rate is None else retry_rate
        out['narrator_gbp'] = turns * direct_turn * (1 + rr) * fx
        out['notes'] = 'Direct cache-aware narrator cost includes retry rate.'
    elif scenario_key == 'C_hosted_gates_mid':
        rr = A['retry_rate_hosted_gate'] if retry_rate is None else retry_rate
        out['narrator_gbp'] = turns * direct_turn * (1 + rr) * fx
        out['gate_gbp'] = turns * gate_turn * fx
        out['fallback_gbp'] = turns * A['api_fallback_rate_hosted_gate'] * direct_turn * fx
        out['notes'] = 'Hosted pre/post gates modelled as 2,200 input + 80 output tokens per turn; 1% narrator API fallback.'
    elif scenario_key == 'D_selfhost_warden':
        rr = A['retry_rate_api'] if retry_rate is None else retry_rate
        out['narrator_gbp'] = turns * direct_turn * (1 + rr) * fx
        count = warden_capacity(mau)
        active = turns * A['warden_gpu_seconds_turn'] / 3600
        paid = count * A['hours_per_month']
        out['gpu_count'] = count
        out['gpu_active_hours'] = active
        out['gpu_paid_hours'] = paid
        out['gpu_idle_hours'] = max(0, paid - active)
        out['gpu_gbp'] = paid * A['warden_gpu_usd_hour'] * fx
        out['fallback_gbp'] = turns * A['api_fallback_rate_warden'] * gate_turn * fx
        out['notes'] = 'Warm RunPod A5000 Pod; self-hosted Warden only. 1% hosted-gate fallback on Warden error/timeout.'
    elif scenario_key == 'E1_selfhost_7b':
        count = capacity(mau, A['full_7b_tps'])
        active = turns * A['output_tokens_turn'] / A['full_7b_tps'] / 3600
        paid = count * A['hours_per_month']
        out['gpu_count'] = count
        out['gpu_active_hours'] = active
        out['gpu_paid_hours'] = paid
        out['gpu_idle_hours'] = max(0, paid - active)
        out['gpu_gbp'] = paid * A['full_7b_gpu_usd_hour'] * fx
        out['fallback_gbp'] = turns * A['api_fallback_rate_full_7b'] * direct_turn * fx
        out['notes'] = '7B is not assumed quality-parity with the paid mid narrator; 85% API fallback is explicitly assumed until evaluation proves otherwise.'
    elif scenario_key == 'E2_selfhost_70b':
        count = capacity(mau, A['full_70b_tps'])
        active = turns * A['output_tokens_turn'] / A['full_70b_tps'] / 3600
        paid = count * A['hours_per_month']
        out['gpu_count'] = count
        out['gpu_active_hours'] = active
        out['gpu_paid_hours'] = paid
        out['gpu_idle_hours'] = max(0, paid - active)
        out['gpu_gbp'] = paid * A['full_70b_gpu_usd_hour'] * fx
        out['fallback_gbp'] = turns * A['api_fallback_rate_full_70b'] * direct_turn * fx
        out['notes'] = '70B capacity is sized for 10x peak and 50 output tok/s/A100; 10% API fallback is explicitly assumed.'
    else:
        raise KeyError(scenario_key)
    out['monthly_total_gbp'] = sum(out[k] for k in ['narrator_gbp', 'gate_gbp', 'gpu_gbp', 'operator_gbp', 'eval_gbp', 'fallback_gbp', 'art_gbp'])
    out['annual_total_gbp'] = out['monthly_total_gbp'] * 12
    out['gbp_per_mau_month'] = out['monthly_total_gbp'] / mau
    return out


def build_results():
    return [compute(key, mau) for key, _, _ in SCENARIOS for mau in A['mau_list']]


def apply_layout(ws, title, subtitle, unit):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 20
    ws.merge_cells('C3:H3')
    c = ws['C3']
    c.value = title
    c.font = Font(color='FFFFFF', bold=True, size=16)
    c.fill = PatternFill('solid', fgColor='135B44')
    c.alignment = Alignment(horizontal='left')
    ws['C5'] = subtitle
    ws['C5'].font = Font(bold=True, size=11)
    ws['C6'] = unit
    ws['C6'].font = Font(italic=True, color='666666')
    ws.freeze_panes = None
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToWidth = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.oddFooter.center.text = '&[Tab] | Page &[Page]'


def comment(cell, text):
    cell.comment = Comment(text, 'Manus AI')


def style_input(cell, source):
    cell.font = Font(color='0000FF')
    cell.alignment = Alignment(horizontal='right')
    comment(cell, source)


def style_formula(cell):
    cell.font = Font(color='000000')
    cell.alignment = Alignment(horizontal='right')


def style_link(cell, source):
    cell.font = Font(color='008000')
    cell.alignment = Alignment(horizontal='right')
    comment(cell, source)


def set_num(cell, fmt, source=None, formula=False):
    cell.number_format = fmt
    if formula:
        style_formula(cell)
    else:
        style_input(cell, source or SOURCE['planning'])


def autofit(ws, start_col=3, max_col=None):
    max_col = max_col or ws.max_column
    for col_idx in range(start_col, max_col + 1):
        letter = get_column_letter(col_idx)
        max_len = 10
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            val = row[0].value
            if val is not None:
                max_len = max(max_len, min(len(str(val)) + 2, 48))
        ws.column_dimensions[letter].width = max_len


def write_assumptions(wb):
    ws = wb.create_sheet('Assumptions')
    apply_layout(ws, 'SynapticGM Own AI vs Paid APIs — Assumptions', 'Editable blue cells drive all calculator outputs; pricing observed 18 Aug 2026.', '($ and £ shown as stated; cost model presents £)')
    sections = [
        ('Workload & policy assumptions', [
            ('MAU low', 100, 'count', SOURCE['planning'], 'mau_low'),
            ('MAU mid', 1000, 'count', SOURCE['planning'], 'mau_mid'),
            ('MAU high', 10000, 'count', SOURCE['planning'], 'mau_high'),
            ('Turns / user / month', A['turns_per_user_month'], 'count', SOURCE['planning'], 'turns_per_user_month'),
            ('Input tokens / turn', A['input_tokens_turn'], 'count', SOURCE['planning'], 'input_tokens_turn'),
            ('Output tokens / turn', A['output_tokens_turn'], 'count', SOURCE['planning'], 'output_tokens_turn'),
            ('Cache hit rate', A['cache_hit_rate'], 'pct', SOURCE['planning'], 'cache_hit_rate'),
            ('Cache-write share if caching', A['cache_write_share_when_enabled'], 'pct', SOURCE['planning'], 'cache_write_share'),
            ('API retry rate', A['retry_rate_api'], 'pct', SOURCE['planning'], 'api_retry_rate'),
            ('Peak factor vs average', A['peak_factor'], 'count', SOURCE['planning'], 'peak_factor'),
        ]),
        ('Currency, people and evaluation', [
            ('USD per GBP', A['usd_per_gbp'], 'fx', SOURCE['usd_per_gbp'], 'usd_per_gbp'),
            ('GBP per USD', A['gbp_per_usd'], 'fx', SOURCE['usd_per_gbp'], 'gbp_per_usd'),
            ('Operator £ / hour', A['operator_gbp_hour'], 'gbp', SOURCE['planning'], 'operator_gbp_hour'),
            ('Days / month', A['days_per_month'], 'count', SOURCE['planning'], 'days_per_month'),
            ('Hours / month for warm GPU', A['hours_per_month'], 'count', SOURCE['planning'], 'hours_per_month'),
            ('Weeks / month', A['weeks_per_month'], 'count', SOURCE['planning'], 'weeks_per_month'),
        ]),
        ('Direct API narrator — Anthropic Sonnet 5 ($ / MTok)', [
            ('Base input', RATES['narrator_input'], 'usd', SOURCE['narrator'], 'narrator_input'),
            ('Cache read', RATES['narrator_cache_read'], 'usd', SOURCE['narrator'], 'narrator_cache_read'),
            ('5m cache write', RATES['narrator_cache_write'], 'usd', SOURCE['narrator'], 'narrator_cache_write'),
            ('Output', RATES['narrator_output'], 'usd', SOURCE['narrator'], 'narrator_output'),
            ('OpenRouter planning loading', RATES['openrouter_loading'], 'pct', SOURCE['openrouter'], 'openrouter_loading'),
        ]),
        ('Hosted low-cost gate — Fireworks Nemotron ($ / MTok)', [
            ('Gate input', RATES['gate_input'], 'usd', SOURCE['gate'], 'gate_input'),
            ('Gate cached input', RATES['gate_cache_read'], 'usd', SOURCE['gate'], 'gate_cache_read'),
            ('Gate output', RATES['gate_output'], 'usd', SOURCE['gate'], 'gate_output'),
            ('Gate input tokens / turn (two passes total)', A['gate_input_tokens_turn'], 'count', SOURCE['planning'], 'gate_input_tokens_turn'),
            ('Gate output tokens / turn (two passes total)', A['gate_output_tokens_turn'], 'count', SOURCE['planning'], 'gate_output_tokens_turn'),
            ('Hosted-gate narrator retry rate', A['retry_rate_hosted_gate'], 'pct', SOURCE['planning'], 'hosted_gate_retry'),
            ('Hosted-gate API fallback rate', A['api_fallback_rate_hosted_gate'], 'pct', SOURCE['planning'], 'hosted_gate_fallback'),
        ]),
        ('Self-host hardware and full-narrator capacity', [
            ('Warden GPU sec / turn (two passes)', A['warden_gpu_seconds_turn'], 'count', SOURCE['planning'], 'warden_gpu_seconds_turn'),
            ('RunPod warm A5000 24GB $ / GPU-hour', A['warden_gpu_usd_hour'], 'usd', SOURCE['runpod_24'], 'warden_gpu_usd_hour'),
            ('RunPod serverless 24GB $ / active GPU-hour', A['warden_serverless_usd_hour'], 'usd', SOURCE['runpod_24'], 'warden_serverless_usd_hour'),
            ('Warden hosted-gate fallback rate', A['api_fallback_rate_warden'], 'pct', SOURCE['planning'], 'warden_fallback_rate'),
            ('7B planning output tokens / second / GPU', A['full_7b_tps'], 'count', SOURCE['planning'], 'full_7b_tps'),
            ('7B A5000 24GB $ / GPU-hour', A['full_7b_gpu_usd_hour'], 'usd', SOURCE['runpod_24'], 'full_7b_gpu_usd_hour'),
            ('7B API fallback / quality-load-shed rate (production-conservative)', A['api_fallback_rate_full_7b'], 'pct', SOURCE['planning'], 'full_7b_fallback'),
            ('70B planning output tokens / second / GPU', A['full_70b_tps'], 'count', SOURCE['planning'], 'full_70b_tps'),
            ('70B A100 PCIe 80GB $ / GPU-hour', A['full_70b_gpu_usd_hour'], 'usd', SOURCE['runpod_80'], 'full_70b_gpu_usd_hour'),
            ('70B API fallback rate', A['api_fallback_rate_full_70b'], 'pct', SOURCE['planning'], 'full_70b_fallback'),
        ]),
        ('Image policy and weekly cap', [
            ('Weekly active share of MAU', A['weekly_active_share'], 'pct', SOURCE['planning'], 'weekly_active_share'),
            ('Image cap / weekly active / week', A['weekly_image_cap_per_wau'], 'count', SOURCE['planning'], 'weekly_image_cap'),
            ('Soft-skip rate', A['art_skip_rate'], 'pct', SOURCE['planning'], 'art_skip_rate'),
            ('FLUX.2 Pro cash cost / first output MP ($)', A['image_usd_per_1mp'], 'usd', SOURCE['image'], 'image_usd_1mp'),
            ('Provider delivery success rate', A['image_delivery_success'], 'pct', SOURCE['planning'], 'image_delivery_success'),
            ('Memorable keep rate after success', A['memorable_keep_rate'], 'pct', SOURCE['planning'], 'memorable_keep_rate'),
        ]),
    ]
    row = 8
    name_cells = {}
    for section, items in sections:
        ws.cell(row=row, column=3, value=section)
        ws.cell(row=row, column=3).fill = PatternFill('solid', fgColor='CFE9E0')
        ws.cell(row=row, column=3).font = Font(bold=True)
        row += 1
        for label, value, fmt, src, key in items:
            ws.cell(row=row, column=3, value=label)
            cell = ws.cell(row=row, column=4, value=value)
            number_format = {'pct': '0.0%', 'usd': '$#,##0.000;($#,##0.000);-', 'gbp': '£#,##0.0;(£#,##0.0);-', 'fx': '0.00000', 'count': '#,##0.0'}[fmt]
            set_num(cell, number_format, src)
            name_cells[key] = cell.coordinate
            row += 1
        row += 1
    # Formula version of GBP per USD as a cross-check, keeping the direct planning input visible as well.
    ws['F8'] = 'Formula check: GBP per USD'
    ws['G8'] = '=1/D21'
    ws['G8'].number_format = '0.00000'
    style_formula(ws['G8'])
    ws['F9'] = 'Model rule'
    ws['G9'] = 'Do not hard-code cells outside this tab.'
    ws['G9'].font = Font(italic=True)
    ws['F10'] = 'Exclusion'
    ws['G10'] = 'Revenue, VAT, taxes, existing platform/ledger cost and fixed corporate overhead are excluded as common/non-incremental.'
    ws['G10'].alignment = Alignment(wrap_text=True)
    autofit(ws)
    ws.print_area = f'B2:{get_column_letter(ws.max_column)}{ws.max_row}'
    return name_cells


def add_cashflow_sheet(wb, cells):
    ws = wb.create_sheet('Cashflow')
    apply_layout(ws, '12-Month Incremental Cashflow — Own AI vs Paid APIs', 'Base case: 40 turns/MAU/month; 30% cache hit; 70% art soft-skip; £0.73784/$ planning FX.', '(£ per month except where stated; annual column is 12× monthly)')
    headers = ['Scenario', 'MAU', 'Turns / month', 'Narrator', 'Hosted gate', 'Warm GPU', 'GPU count', 'GPU active h', 'GPU paid h', 'GPU idle h', 'Operator', 'Eval', 'Failure fallback', 'Art', 'Total / month', 'Total / 12m', '£ / MAU / month']
    start_row = 8
    for col, header in enumerate(headers, 3):
        c = ws.cell(start_row, col, header)
        c.fill = PatternFill('solid', fgColor='CFE9E0')
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal='center', wrap_text=True)
    results = build_results()
    # Inputs from Assumptions are linked cross-sheet (green); cashflow columns use formulas.
    scenario_rows = []
    row = start_row + 1
    for s_idx, (key, label, desc) in enumerate(SCENARIOS):
        for mau_idx in range(3):
            r = results[s_idx * 3 + mau_idx]
            mau_cell = [cells['mau_low'], cells['mau_mid'], cells['mau_high']][mau_idx]
            ws.cell(row, 3, label)
            ws.cell(row, 4, f'=Assumptions!${mau_cell[0]}${mau_cell[1:]}')
            style_link(ws.cell(row, 4), SOURCE['planning'])
            ws.cell(row, 5, f'=D{row}*Assumptions!${cells["turns_per_user_month"][0]}${cells["turns_per_user_month"][1:]}')
            style_formula(ws.cell(row, 5))
            # Formula setup encoded as readable formula per scenario; static computed amounts are only for validation exported separately.
            fx = f'Assumptions!${cells["gbp_per_usd"][0]}${cells["gbp_per_usd"][1:]}'
            days = f'Assumptions!${cells["days_per_month"][0]}${cells["days_per_month"][1:]}'
            turns = f'E{row}'
            hit = f'Assumptions!${cells["cache_hit_rate"][0]}${cells["cache_hit_rate"][1:]}'
            write = f'Assumptions!${cells["cache_write_share"][0]}${cells["cache_write_share"][1:]}'
            ntin = f'Assumptions!${cells["narrator_input"][0]}${cells["narrator_input"][1:]}'
            ntcr = f'Assumptions!${cells["narrator_cache_read"][0]}${cells["narrator_cache_read"][1:]}'
            ntcw = f'Assumptions!${cells["narrator_cache_write"][0]}${cells["narrator_cache_write"][1:]}'
            ntout = f'Assumptions!${cells["narrator_output"][0]}${cells["narrator_output"][1:]}'
            in_tok = f'Assumptions!${cells["input_tokens_turn"][0]}${cells["input_tokens_turn"][1:]}'
            out_tok = f'Assumptions!${cells["output_tokens_turn"][0]}${cells["output_tokens_turn"][1:]}'
            direct_turn = f'(({in_tok}/1000000)*({hit}*{ntcr}+IF({hit}>0,{write}*{ntcw},0)+(1-{hit}-IF({hit}>0,{write},0))*{ntin})+({out_tok}/1000000)*{ntout})'
            openrouter_turn = f'({direct_turn}*(1+Assumptions!${cells["openrouter_loading"][0]}${cells["openrouter_loading"][1:]}))'
            gate_turn = f'((Assumptions!${cells["gate_input_tokens_turn"][0]}${cells["gate_input_tokens_turn"][1:]}/1000000)*({hit}*Assumptions!${cells["gate_cache_read"][0]}${cells["gate_cache_read"][1:]}+(1-{hit})*Assumptions!${cells["gate_input"][0]}${cells["gate_input"][1:]})+(Assumptions!${cells["gate_output_tokens_turn"][0]}${cells["gate_output_tokens_turn"][1:]}/1000000)*Assumptions!${cells["gate_output"][0]}${cells["gate_output"][1:]})'
            # Columns: F narrator, G gate, H GPU, I count, J active, K paid, L idle, M ops, N eval, O fallback, P art, Q monthly, R annual, S per MAU.
            if key == 'A_openrouter_only':
                ws.cell(row, 6, f'={turns}*{openrouter_turn}*(1+Assumptions!${cells["api_retry_rate"][0]}${cells["api_retry_rate"][1:]})*{fx}')
                values = {'g': 0, 'h': 0, 'i': 0, 'j': 0, 'k': 0, 'l': 0, 'o': 0, 'ops_h': OPS[key]['hours'], 'eval': OPS[key]['eval_gbp']}
            elif key == 'B_direct_cache':
                ws.cell(row, 6, f'={turns}*{direct_turn}*(1+Assumptions!${cells["api_retry_rate"][0]}${cells["api_retry_rate"][1:]})*{fx}')
                values = {'g': 0, 'h': 0, 'i': 0, 'j': 0, 'k': 0, 'l': 0, 'o': 0, 'ops_h': OPS[key]['hours'], 'eval': OPS[key]['eval_gbp']}
            elif key == 'C_hosted_gates_mid':
                ws.cell(row, 6, f'={turns}*{direct_turn}*(1+Assumptions!${cells["hosted_gate_retry"][0]}${cells["hosted_gate_retry"][1:]})*{fx}')
                ws.cell(row, 7, f'={turns}*{gate_turn}*{fx}')
                ws.cell(row, 15, f'={turns}*Assumptions!${cells["hosted_gate_fallback"][0]}${cells["hosted_gate_fallback"][1:]}*{direct_turn}*{fx}')
                values = {'h': 0, 'i': 0, 'j': 0, 'k': 0, 'l': 0, 'ops_h': OPS[key]['hours'], 'eval': OPS[key]['eval_gbp']}
            elif key == 'D_selfhost_warden':
                ws.cell(row, 6, f'={turns}*{direct_turn}*(1+Assumptions!${cells["api_retry_rate"][0]}${cells["api_retry_rate"][1:]})*{fx}')
                ws.cell(row, 9, f'=MAX(1,ROUNDUP(({turns}/({days}*24*3600))*Assumptions!${cells["peak_factor"][0]}${cells["peak_factor"][1:]}*Assumptions!${cells["warden_gpu_seconds_turn"][0]}${cells["warden_gpu_seconds_turn"][1:]},0))')
                ws.cell(row, 10, f'={turns}*Assumptions!${cells["warden_gpu_seconds_turn"][0]}${cells["warden_gpu_seconds_turn"][1:]}/3600')
                ws.cell(row, 11, f'=I{row}*Assumptions!${cells["hours_per_month"][0]}${cells["hours_per_month"][1:]}')
                ws.cell(row, 12, f'=MAX(0,K{row}-J{row})')
                ws.cell(row, 8, f'=K{row}*Assumptions!${cells["warden_gpu_usd_hour"][0]}${cells["warden_gpu_usd_hour"][1:]}*{fx}')
                ws.cell(row, 15, f'={turns}*Assumptions!${cells["warden_fallback_rate"][0]}${cells["warden_fallback_rate"][1:]}*{gate_turn}*{fx}')
                values = {'g': 0, 'ops_h': OPS[key]['hours'], 'eval': OPS[key]['eval_gbp']}
            elif key == 'E1_selfhost_7b':
                ws.cell(row, 9, f'=MAX(1,ROUNDUP((({turns}/({days}*24*3600))*Assumptions!${cells["peak_factor"][0]}${cells["peak_factor"][1:]}*Assumptions!${cells["output_tokens_turn"][0]}${cells["output_tokens_turn"][1:]})/Assumptions!${cells["full_7b_tps"][0]}${cells["full_7b_tps"][1:]},0))')
                ws.cell(row, 10, f'={turns}*Assumptions!${cells["output_tokens_turn"][0]}${cells["output_tokens_turn"][1:]}/Assumptions!${cells["full_7b_tps"][0]}${cells["full_7b_tps"][1:]}/3600')
                ws.cell(row, 11, f'=I{row}*Assumptions!${cells["hours_per_month"][0]}${cells["hours_per_month"][1:]}')
                ws.cell(row, 12, f'=MAX(0,K{row}-J{row})')
                ws.cell(row, 8, f'=K{row}*Assumptions!${cells["full_7b_gpu_usd_hour"][0]}${cells["full_7b_gpu_usd_hour"][1:]}*{fx}')
                ws.cell(row, 15, f'={turns}*Assumptions!${cells["full_7b_fallback"][0]}${cells["full_7b_fallback"][1:]}*{direct_turn}*{fx}')
                values = {'g': 0, 'ops_h': OPS[key]['hours'], 'eval': OPS[key]['eval_gbp']}
            elif key == 'E2_selfhost_70b':
                ws.cell(row, 9, f'=MAX(1,ROUNDUP((({turns}/({days}*24*3600))*Assumptions!${cells["peak_factor"][0]}${cells["peak_factor"][1:]}*Assumptions!${cells["output_tokens_turn"][0]}${cells["output_tokens_turn"][1:]})/Assumptions!${cells["full_70b_tps"][0]}${cells["full_70b_tps"][1:]},0))')
                ws.cell(row, 10, f'={turns}*Assumptions!${cells["output_tokens_turn"][0]}${cells["output_tokens_turn"][1:]}/Assumptions!${cells["full_70b_tps"][0]}${cells["full_70b_tps"][1:]}/3600')
                ws.cell(row, 11, f'=I{row}*Assumptions!${cells["hours_per_month"][0]}${cells["hours_per_month"][1:]}')
                ws.cell(row, 12, f'=MAX(0,K{row}-J{row})')
                ws.cell(row, 8, f'=K{row}*Assumptions!${cells["full_70b_gpu_usd_hour"][0]}${cells["full_70b_gpu_usd_hour"][1:]}*{fx}')
                ws.cell(row, 15, f'={turns}*Assumptions!${cells["full_70b_fallback"][0]}${cells["full_70b_fallback"][1:]}*{direct_turn}*{fx}')
                values = {'g': 0, 'ops_h': OPS[key]['hours'], 'eval': OPS[key]['eval_gbp']}
            # Fill blank numerical regions with zero and style formula cells.
            for col in range(6, 16):
                c = ws.cell(row, col)
                if c.value is None:
                    c.value = 0
                if isinstance(c.value, str) and c.value.startswith('='):
                    style_formula(c)
                else:
                    style_formula(c)
            ws.cell(row, 13, f'={values["ops_h"]}*Assumptions!${cells["operator_gbp_hour"][0]}${cells["operator_gbp_hour"][1:]}')
            ws.cell(row, 14, f'={values["eval"]}')
            ws.cell(row, 16, f'=D{row}*Assumptions!${cells["weekly_active_share"][0]}${cells["weekly_active_share"][1:]}*Assumptions!${cells["weekly_image_cap"][0]}${cells["weekly_image_cap"][1:]}*Assumptions!${cells["weeks_per_month"][0]}${cells["weeks_per_month"][1:]}*(1-Assumptions!${cells["art_skip_rate"][0]}${cells["art_skip_rate"][1:]})*Assumptions!${cells["image_usd_1mp"][0]}${cells["image_usd_1mp"][1:]}*{fx}')
            ws.cell(row, 17, f'=SUM(F{row}:H{row},M{row}:P{row})')
            ws.cell(row, 18, f'=Q{row}*12')
            ws.cell(row, 19, f'=Q{row}/D{row}')
            for col in [13, 14, 16, 17, 18, 19]:
                style_formula(ws.cell(row, col))
            for col in [6,7,8,13,14,15,16,17,18,19]:
                ws.cell(row, col).number_format = '£#,##0.0;(£#,##0.0);-'
            for col in [9,10,11,12]:
                ws.cell(row, col).number_format = '#,##0.0'
            ws.cell(row, 3).alignment = Alignment(wrap_text=True)
            scenario_rows.append((key, row))
            row += 1
        row += 1
    # Notes & control descriptions.
    note_row = row + 1
    ws.cell(note_row, 3, 'Interpretation controls').fill = PatternFill('solid', fgColor='CFE9E0')
    ws.cell(note_row, 3).font = Font(bold=True)
    notes = [
        'A is a planning OpenRouter gross-up, not a claim of universal transaction fees. Model-route prices vary by provider.',
        'B–D use the same direct Sonnet 5 mid-narrator workload to isolate routing/gate/warden economics.',
        'D models a warm production pod to expose idle cost. A scale-to-zero serverless Warden is reported separately in Sensitivity.',
        'E1 assumes 85% paid-API fallback because a 7B narrator is not granted quality parity without evaluation proof; E2 is a quality-aspirational 70B capacity case. Neither is asserted to match paid narrator quality without benchmark proof.',
        'GPU idle hours are paid warm hours minus estimated active compute hours. Active hours do not include cold starts, storage, egress or monitoring overhead.',
        'Art is a common direct BFL baseline at a planning cap of one attempt per weekly-active user per week, 60% WAU/MAU, with a 70% soft skip rate.',
    ]
    for note in notes:
        note_row += 1
        ws.cell(note_row, 3, '• ' + note)
        ws.merge_cells(start_row=note_row, start_column=3, end_row=note_row, end_column=19)
        ws.cell(note_row, 3).alignment = Alignment(wrap_text=True, vertical='top')
    # Conditional formatting emphasizes high idle paid time.
    ws.conditional_formatting.add(f'L{start_row+1}:L{row-2}', CellIsRule(operator='greaterThan', formula=['500'], fill=PatternFill('solid', fgColor='FCE4D6')))
    # Build a bar chart (monthly cost by MAU; separate data subset placed out of print area).
    chart_start = note_row + 3
    ws.cell(chart_start, 3, 'Chart helper: monthly £ by scenario at 1,000 MAU')
    for ci, head in enumerate(['Scenario', 'Monthly £'], 3):
        ws.cell(chart_start + 1, ci, head)
    rr = chart_start + 2
    for key, label, _ in SCENARIOS:
        found = next(r for k, r in scenario_rows if k == key and ws.cell(r, 4).value == '=Assumptions!$D$10') if False else None
        # Use the static row map: each scenario second MAU row is start + 1 within its three-row block.
    for idx, (key, label, _) in enumerate(SCENARIOS):
        source_row = start_row + 1 + idx * 4 + 1
        ws.cell(rr + idx, 3, label)
        ws.cell(rr + idx, 4, f'=Q{source_row}')
        style_formula(ws.cell(rr + idx, 4))
        ws.cell(rr + idx, 4).number_format = '£#,##0.0;(£#,##0.0);-'
    chart = BarChart()
    chart.type = 'bar'
    chart.style = 10
    chart.title = 'Monthly Cost at 1,000 MAU'
    chart.y_axis.title = 'Scenario'
    chart.x_axis.title = 'GBP/month'
    data = Reference(ws, min_col=4, min_row=chart_start + 1, max_row=chart_start + 1 + len(SCENARIOS))
    cats = Reference(ws, min_col=3, min_row=chart_start + 2, max_row=chart_start + 1 + len(SCENARIOS))
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 9
    chart.width = 17
    ws.add_chart(chart, f'C{chart_start}')
    autofit(ws, max_col=19)
    ws.print_area = f'B2:S{note_row}'


def add_sensitivity_sheet(wb, cells):
    ws = wb.create_sheet('Sensitivity')
    apply_layout(ws, 'Break-Even Calculator and Sensitivities', 'Every denominator and workload variable is editable on Assumptions. Use this tab to stress cache, retry, art skip, and Warden deployment mode.', '(£ unless stated)')
    r = 8
    ws.cell(r, 3, 'Core spreadsheet-friendly formulas').fill = PatternFill('solid', fgColor='CFE9E0')
    ws.cell(r, 3).font = Font(bold=True)
    formulas = [
        ('Monthly turns', 'MAU × turns/user/month'),
        ('Narrator $/turn', '(Input tokens ÷ 1,000,000 × [H×cache-read + W×cache-write + (1−H−W)×base-input]) + (Output tokens ÷ 1,000,000 × output rate)'),
        ('Narrator £/month', 'Monthly turns × narrator $/turn × (1 + retry rate) × GBP/USD'),
        ('Art £/month', 'MAU × WAU share × weekly cap × weeks/month × (1−art skip) × $/image × GBP/USD'),
        ('Warm GPU £/month', 'GPU count × warm hours/month × $/GPU-hour × GBP/USD'),
        ('Warden GPU count', 'MAX(1, CEILING(avg turns/sec × peak factor × GPU seconds/turn, 1))'),
        ('Full narrator GPU count', 'MAX(1, CEILING(peak turns/sec × output tokens/turn ÷ output tokens/sec/GPU, 1))'),
        ('Break-even MAU vs paid API', '(self-host fixed £/month − paid fixed £/month) ÷ (paid variable £/MAU/month − self-host variable £/MAU/month)')
    ]
    r += 1
    for label, formula in formulas:
        ws.cell(r, 3, label).font = Font(bold=True)
        ws.cell(r, 4, formula)
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=10)
        ws.cell(r, 4).alignment = Alignment(wrap_text=True)
        r += 1
    r += 1
    ws.cell(r, 3, 'Cache-hit sensitivity — Direct APIs + prompt cache').fill = PatternFill('solid', fgColor='CFE9E0')
    ws.cell(r, 3).font = Font(bold=True)
    r += 1
    headers = ['Cache hit', '100 MAU monthly £', '1,000 MAU monthly £', '10,000 MAU monthly £', 'Narrator $/turn']
    for i, h in enumerate(headers, 3):
        ws.cell(r, i, h).font = Font(bold=True)
        ws.cell(r, i).fill = PatternFill('solid', fgColor='CFE9E0')
    cache_rows = [0.0, 0.30, 0.70]
    for h in cache_rows:
        r += 1
        ws.cell(r, 3, h)
        set_num(ws.cell(r, 3), '0.0%', SOURCE['planning'])
        # Static values are formula-derived through visible Excel expression using explicit stress input value in C.
        for idx, mau_key in enumerate(['mau_low', 'mau_mid', 'mau_high'], 4):
            mau_ref = f'Assumptions!${cells[mau_key][0]}${cells[mau_key][1:]}'
            in_tok = f'Assumptions!${cells["input_tokens_turn"][0]}${cells["input_tokens_turn"][1:]}'
            out_tok = f'Assumptions!${cells["output_tokens_turn"][0]}${cells["output_tokens_turn"][1:]}'
            ni = f'Assumptions!${cells["narrator_input"][0]}${cells["narrator_input"][1:]}'
            cr = f'Assumptions!${cells["narrator_cache_read"][0]}${cells["narrator_cache_read"][1:]}'
            cw = f'Assumptions!${cells["narrator_cache_write"][0]}${cells["narrator_cache_write"][1:]}'
            no = f'Assumptions!${cells["narrator_output"][0]}${cells["narrator_output"][1:]}'
            write = f'Assumptions!${cells["cache_write_share"][0]}${cells["cache_write_share"][1:]}'
            rr = f'Assumptions!${cells["api_retry_rate"][0]}${cells["api_retry_rate"][1:]}'
            fx = f'Assumptions!${cells["gbp_per_usd"][0]}${cells["gbp_per_usd"][1:]}'
            art = f'({mau_ref}*Assumptions!${cells["weekly_active_share"][0]}${cells["weekly_active_share"][1:]}*Assumptions!${cells["weekly_image_cap"][0]}${cells["weekly_image_cap"][1:]}*Assumptions!${cells["weeks_per_month"][0]}${cells["weeks_per_month"][1:]}*(1-Assumptions!${cells["art_skip_rate"][0]}${cells["art_skip_rate"][1:]})*Assumptions!${cells["image_usd_1mp"][0]}${cells["image_usd_1mp"][1:]}*{fx})'
            perturn = f'(({in_tok}/1000000)*($C{r}*{cr}+IF($C{r}>0,{write}*{cw},0)+(1-$C{r}-IF($C{r}>0,{write},0))*{ni})+({out_tok}/1000000)*{no})'
            ws.cell(r, idx, f'={mau_ref}*Assumptions!${cells["turns_per_user_month"][0]}${cells["turns_per_user_month"][1:]}*{perturn}*(1+{rr})*{fx}+{art}+2*Assumptions!${cells["operator_gbp_hour"][0]}${cells["operator_gbp_hour"][1:]}+40')
            ws.cell(r, idx).number_format = '£#,##0.0;(£#,##0.0);-'
            style_formula(ws.cell(r, idx))
        ws.cell(r, 8, f'={perturn}')
        ws.cell(r, 8).number_format = '$0.000000'
        style_formula(ws.cell(r, 8))
    r += 2
    ws.cell(r, 3, 'Retry-rate sensitivity — Direct APIs + prompt cache at 1,000 MAU').fill = PatternFill('solid', fgColor='CFE9E0')
    ws.cell(r, 3).font = Font(bold=True)
    r += 1
    for i, h in enumerate(['Retry rate', 'Narrator £/month', 'Total £/month', 'Incremental vs 0%'], 3):
        ws.cell(r, i, h).font = Font(bold=True)
        ws.cell(r, i).fill = PatternFill('solid', fgColor='CFE9E0')
    retry_start = r + 1
    for rr_val in [0.0, 0.05, 0.15]:
        r += 1
        ws.cell(r, 3, rr_val)
        set_num(ws.cell(r, 3), '0.0%', SOURCE['planning'])
        mau_ref = f'Assumptions!${cells["mau_mid"][0]}${cells["mau_mid"][1:]}'
        in_tok = f'Assumptions!${cells["input_tokens_turn"][0]}${cells["input_tokens_turn"][1:]}'
        out_tok = f'Assumptions!${cells["output_tokens_turn"][0]}${cells["output_tokens_turn"][1:]}'
        hit = f'Assumptions!${cells["cache_hit_rate"][0]}${cells["cache_hit_rate"][1:]}'
        write = f'Assumptions!${cells["cache_write_share"][0]}${cells["cache_write_share"][1:]}'
        ni = f'Assumptions!${cells["narrator_input"][0]}${cells["narrator_input"][1:]}'
        cr = f'Assumptions!${cells["narrator_cache_read"][0]}${cells["narrator_cache_read"][1:]}'
        cw = f'Assumptions!${cells["narrator_cache_write"][0]}${cells["narrator_cache_write"][1:]}'
        no = f'Assumptions!${cells["narrator_output"][0]}${cells["narrator_output"][1:]}'
        fx = f'Assumptions!${cells["gbp_per_usd"][0]}${cells["gbp_per_usd"][1:]}'
        direct_perturn = f'(({in_tok}/1000000)*({hit}*{cr}+IF({hit}>0,{write}*{cw},0)+(1-{hit}-IF({hit}>0,{write},0))*{ni})+({out_tok}/1000000)*{no})'
        art = f'({mau_ref}*Assumptions!${cells["weekly_active_share"][0]}${cells["weekly_active_share"][1:]}*Assumptions!${cells["weekly_image_cap"][0]}${cells["weekly_image_cap"][1:]}*Assumptions!${cells["weeks_per_month"][0]}${cells["weeks_per_month"][1:]}*(1-Assumptions!${cells["art_skip_rate"][0]}${cells["art_skip_rate"][1:]})*Assumptions!${cells["image_usd_1mp"][0]}${cells["image_usd_1mp"][1:]}*{fx})'
        ws.cell(r, 4, f'={mau_ref}*Assumptions!${cells["turns_per_user_month"][0]}${cells["turns_per_user_month"][1:]}*{direct_perturn}*(1+$C{r})*{fx}')
        ws.cell(r, 4).number_format = '£#,##0.0;(£#,##0.0);-'
        style_formula(ws.cell(r, 4))
        ws.cell(r, 5, f'=D{r}+{art}+2*Assumptions!${cells["operator_gbp_hour"][0]}${cells["operator_gbp_hour"][1:]}+40')
        ws.cell(r, 5).number_format = '£#,##0.0;(£#,##0.0);-'
        style_formula(ws.cell(r, 5))
        ws.cell(r, 6, f'=E{r}-$E${retry_start}')
        ws.cell(r, 6).number_format = '£#,##0.0;(£#,##0.0);-'
        style_formula(ws.cell(r, 6))
    r += 2
    ws.cell(r, 3, 'Art soft-skip sensitivity — 1 MP BFL/FLUX.2 Pro cash spend').fill = PatternFill('solid', fgColor='CFE9E0')
    ws.cell(r, 3).font = Font(bold=True)
    r += 1
    for i, h in enumerate(['Art skip rate', '100 MAU £/month', '1,000 MAU £/month', '10,000 MAU £/month', 'Image attempts / 10k MAU / month'], 3):
        ws.cell(r, i, h).font = Font(bold=True)
        ws.cell(r, i).fill = PatternFill('solid', fgColor='CFE9E0')
    for skip in [0.0, 0.30, 0.70]:
        r += 1
        ws.cell(r, 3, skip)
        set_num(ws.cell(r, 3), '0.0%', SOURCE['planning'])
        for idx, mau_key in enumerate(['mau_low', 'mau_mid', 'mau_high'], 4):
            mau_ref = f'Assumptions!${cells[mau_key][0]}${cells[mau_key][1:]}'
            formula = f'={mau_ref}*Assumptions!${cells["weekly_active_share"][0]}${cells["weekly_active_share"][1:]}*Assumptions!${cells["weekly_image_cap"][0]}${cells["weekly_image_cap"][1:]}*Assumptions!${cells["weeks_per_month"][0]}${cells["weeks_per_month"][1:]}*(1-$C{r})*Assumptions!${cells["image_usd_1mp"][0]}${cells["image_usd_1mp"][1:]}*Assumptions!${cells["gbp_per_usd"][0]}${cells["gbp_per_usd"][1:]}'
            ws.cell(r, idx, formula)
            ws.cell(r, idx).number_format = '£#,##0.0;(£#,##0.0);-'
            style_formula(ws.cell(r, idx))
        mau_ref = f'Assumptions!${cells["mau_high"][0]}${cells["mau_high"][1:]}'
        attempts = f'={mau_ref}*Assumptions!${cells["weekly_active_share"][0]}${cells["weekly_active_share"][1:]}*Assumptions!${cells["weekly_image_cap"][0]}${cells["weekly_image_cap"][1:]}*Assumptions!${cells["weeks_per_month"][0]}${cells["weeks_per_month"][1:]}*(1-$C{r})'
        ws.cell(r, 8, attempts)
        ws.cell(r, 8).number_format = '#,##0.0'
        style_formula(ws.cell(r, 8))
    r += 2
    ws.cell(r, 3, 'Warden deployment comparison — 24GB worker at baseline active seconds').fill = PatternFill('solid', fgColor='CFE9E0')
    ws.cell(r, 3).font = Font(bold=True)
    r += 1
    for i, h in enumerate(['MAU', 'Active GPU h/month', 'Warm Pod £/month', 'Serverless active-only £/month', 'Warm idle h/month', 'Cold-start caveat'], 3):
        ws.cell(r, i, h).font = Font(bold=True)
        ws.cell(r, i).fill = PatternFill('solid', fgColor='CFE9E0')
    for mau in A['mau_list']:
        r += 1
        active = mau * A['turns_per_user_month'] * A['warden_gpu_seconds_turn'] / 3600
        paid = A['hours_per_month']
        ws.cell(r, 3, mau); set_num(ws.cell(r, 3), '#,##0', SOURCE['planning'])
        ws.cell(r, 4, active); set_num(ws.cell(r, 4), '#,##0.0', SOURCE['planning'])
        ws.cell(r, 5, paid * A['warden_gpu_usd_hour'] * A['gbp_per_usd']); set_num(ws.cell(r, 5), '£#,##0.0;(£#,##0.0);-', SOURCE['runpod_24'])
        ws.cell(r, 6, active * A['warden_serverless_usd_hour'] * A['gbp_per_usd']); set_num(ws.cell(r, 6), '£#,##0.0;(£#,##0.0);-', SOURCE['runpod_24'])
        ws.cell(r, 7, max(0, paid - active)); set_num(ws.cell(r, 7), '#,##0.0', SOURCE['planning'])
        ws.cell(r, 8, 'Serverless removes warm idle cost but can introduce start/queue latency; validate against live turn SLO before relying on it.')
        ws.cell(r, 8).alignment = Alignment(wrap_text=True)
    autofit(ws, max_col=10)
    ws.print_area = f'B2:J{r}'


def add_image_sheet(wb, cells):
    ws = wb.create_sheet('Image Ladder')
    apply_layout(ws, 'Image Cost Ladder — FLUX Options for SynapticGM', 'Per 1 MP baseline, with filterability assessed as implementation controls rather than a guarantee of child suitability.', '(USD list price, £ planning translation, and cost per retained memorable image)')
    gbp = f'Assumptions!${cells["gbp_per_usd"][0]}${cells["gbp_per_usd"][1:]}'
    delivery = f'Assumptions!${cells["image_delivery_success"][0]}${cells["image_delivery_success"][1:]}'
    keep = f'Assumptions!${cells["memorable_keep_rate"][0]}${cells["memorable_keep_rate"][1:]}'
    headers = ['Route', '1 MP list price $', '1 MP £', 'Kid/adult filterability', 'Cost per retained memorable £', 'Live implication', 'Source / access date']
    row = 8
    for i, h in enumerate(headers, 3):
        ws.cell(row, i, h).font = Font(bold=True)
        ws.cell(row, i).fill = PatternFill('solid', fgColor='CFE9E0')
        ws.cell(row, i).alignment = Alignment(horizontal='center', wrap_text=True)
    ladder = [
        ('FLUX via OpenRouter — FLUX.2 Pro', 0.030, 'Provider-policy and SynapticGM gate required; model page exposes price, but no complete child-policy control was evidenced.', 'No list-price advantage over direct BFL at 1 MP. Use only if unified routing/logging is worth the account-level loading.', 'OpenRouter FLUX.2 Pro, accessed 18 Aug 2026: https://openrouter.ai/black-forest-labs/flux.2-pro'),
        ('BFL direct — FLUX.2 Pro', 0.030, 'BFL exposes `safety_tolerance` and moderation statuses; usage policy bars harmful/minor sexual content. SynapticGM remains policy owner.', 'Cleanest direct relationship; async API; retrieve/re-host signed result rather than expose provider URL.', 'BFL Pricing / Usage Policy, accessed 18 Aug 2026: https://docs.bfl.ai/quick_start/pricing ; https://bfl.ai/legal/usage-policy'),
        ('Replicate — FLUX.2 Pro', 0.030, 'Safety checker exists for Flux base / derivative fine-tunes and can be disabled for API custom safety. Configurable, but external kid policy required.', '1 MP no-reference baseline combines $0.015/run + $0.015 output MP; reference images add $0.015/MP.', 'Replicate FLUX / Safety Checking, accessed 18 Aug 2026: https://replicate.com/black-forest-labs/flux-2-pro ; https://replicate.com/docs/topics/predictions/safety-checking'),
        ('fal — FLUX.2 Pro', 0.030, '`enable_safety_checker` defaults true; blocked outputs are black and `has_nsfw_concepts` is reported. This is NSFW control, not complete child context policy.', 'Default-on provider safety knob is valuable defense-in-depth; add SynapticGM prompt+image policy before display.', 'fal FLUX.2 Pro/API, accessed 18 Aug 2026: https://fal.ai/models/fal-ai/flux-2-pro ; https://fal.ai/models/fal-ai/flux-2/api'),
        ('Local ComfyUI — FLUX/open model', None, 'Maximum control only if SynapticGM selects, tests, versions and operates local filters; no platform-wide universal filter.', 'Not zero. Active GPU £ + operator + safety/eval burden; calculate locally from workflow seconds and GPU quote.', 'ComfyUI repository / Desktop docs, accessed 18 Aug 2026: https://github.com/Comfy-Org/ComfyUI ; https://docs.comfy.org/installation/desktop/overview'),
    ]
    for route, usd, filterability, implication, src in ladder:
        row += 1
        ws.cell(row, 3, route)
        if usd is not None:
            ws.cell(row, 4, usd); set_num(ws.cell(row, 4), '$0.000', SOURCE['image'])
            ws.cell(row, 5, f'=D{row}*{gbp}')
            ws.cell(row, 5).number_format = '£0.000'
            style_formula(ws.cell(row, 5))
            ws.cell(row, 7, f'=E{row}/({delivery}*{keep})')
            ws.cell(row, 7).number_format = '£0.000'
            style_formula(ws.cell(row, 7))
        else:
            ws.cell(row, 4, 'GPU dependent')
            ws.cell(row, 5, 'GPU dependent')
            ws.cell(row, 7, 'GPU/operator dependent')
        ws.cell(row, 6, filterability)
        ws.cell(row, 8, implication)
        ws.cell(row, 9, src)
        for c in range(3, 10):
            ws.cell(row, c).alignment = Alignment(wrap_text=True, vertical='top')
    row += 2
    ws.cell(row, 3, 'Soft-skip policy').fill = PatternFill('solid', fgColor='CFE9E0')
    ws.cell(row, 3).font = Font(bold=True)
    row += 1
    policy = [
        ('Default trigger', 'Do not generate merely because a turn exists. Generate only when a ledger-supported, player-visible, distinct moment passes the stated weekly cap and the prompt/image policy gate.'),
        ('Default cap', 'One eligible image attempt per weekly active user per week, 60% WAU/MAU planning assumption. The workbook exposes 0%, 30%, and 70% soft-skip stress cases.'),
        ('Default skip', '70% of cap-eligible image opportunities are soft-skipped to preserve the feeling of a memorable moment rather than turn-by-turn wallpaper.'),
        ('Paid output rule', 'Assume a submitted image request can incur its listed fee. Do not assume provider-moderated, failed, duplicate, or rejected outputs are free unless billing documentation proves it.'),
        ('Display rule', 'Run SynapticGM prompt and result policy before display. Treat provider NSFW/moderation as an additional layer, not the complete kid/adult decision.'),
        ('Metric', 'Cash per retained memorable image = paid cost / (provider delivery success × memorable keep rate). Baseline uses 90% delivery success and 65% keep rate; change both in Assumptions.'),
    ]
    for title, body in policy:
        ws.cell(row, 3, title).font = Font(bold=True)
        ws.cell(row, 4, body)
        ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=9)
        ws.cell(row, 4).alignment = Alignment(wrap_text=True, vertical='top')
        row += 1
    autofit(ws, max_col=9)
    ws.print_area = f'B2:I{row}'


def add_sources_sheet(wb):
    ws = wb.create_sheet('Sources')
    apply_layout(ws, 'Fresh Price Sources and Model Caveats', 'All vendor price pages accessed 18 Aug 2026. List prices may change; marketplace prices are snapshots, not commitments.', '(source log)')
    headers = ['Category', 'Vendor / source', 'Observed item used', 'URL', 'Access date', 'Caveat']
    row = 8
    for i, h in enumerate(headers, 3):
        ws.cell(row, i, h).font = Font(bold=True)
        ws.cell(row, i).fill = PatternFill('solid', fgColor='CFE9E0')
    records = [
        ('Text narrator', 'Anthropic', 'Claude Sonnet 5: $2 input / $0.20 cache read / $2.50 5m write / $10 output per MTok', 'https://docs.anthropic.com/en/docs/about-claude/pricing', '2026-08-18', 'Direct API list price; cache writes and reads are distinct charges.'),
        ('Routing', 'OpenRouter', 'Pay-as-you-go platform fee shown as 5.5%; sticky prompt caching documentation', 'https://openrouter.ai/pricing ; https://openrouter.ai/docs/guides/best-practices/prompt-caching', '2026-08-18', 'Planning gross-up; model/provider rate and billing treatment must be verified in account.'),
        ('Hosted gate', 'Fireworks', 'Nemotron 3.5 Lightning 30B A3B: $0.05 / $0.01 cached / $0.20 output per MTok', 'https://docs.fireworks.ai/serverless/pricing', '2026-08-18', 'Model-specific serverless list price; no warm idle charge assumed.'),
        ('Direct low-cost reference', 'DeepSeek', 'V4 Flash off-peak: $0.007 cache hit / $0.22 miss / $0.66 output per MTok', 'https://api-docs.deepseek.com/quick_start/pricing/', '2026-08-18', 'Peak/off-peak varies; not used as the base mid-narrator.'),
        ('24GB GPU', 'RunPod', 'A5000 Pod $0.27/h; 24GB serverless group $0.69/h; 4090 Pod $0.74/h', 'https://www.runpod.io/pricing', '2026-08-18', 'Public on-demand quote; warm pod assumes 730 paid h/month.'),
        ('80GB GPU', 'RunPod', 'A100 PCIe 80GB Pod $1.39/h; A100 SXM 80GB $1.59/h', 'https://www.runpod.io/pricing', '2026-08-18', 'Used for 70B capacity illustration; hardware fit/performance requires benchmark validation.'),
        ('GPU comparison', 'Lambda', 'A100 40GB $1.99/h; A100 80GB $2.79/h; no current public 24/48GB row observed', 'https://lambda.ai/pricing', '2026-08-18', 'Do not fabricate a current 24GB Lambda quote.'),
        ('GPU comparison', 'Vast.ai', 'Marketplace snapshot captured: RTX 3090 24GB $0.07/h; RTX 4090 24GB $0.13/h', 'https://vast.ai/pricing ; https://vast.ai/pricing/gpu/RTX-4090', '2026-08-18', 'Marketplace offer, host/reliability/storage/bandwidth variable; do not use as production commitment baseline.'),
        ('Image', 'BFL', 'FLUX.2 Pro T2I from $0.03 first MP; 1 credit = $0.01', 'https://docs.bfl.ai/quick_start/pricing', '2026-08-18', 'Resolution and input-editing dependent.'),
        ('Image', 'OpenRouter', 'FLUX.2 Pro $0.03 first output MP; $0.015 input MP for reference images', 'https://openrouter.ai/black-forest-labs/flux.2-pro', '2026-08-18', 'No assumed price advantage over direct BFL at 1MP.'),
        ('Image', 'Replicate', 'FLUX.2 Pro $0.015/run plus $0.015 per output/input MP in source sweep', 'https://replicate.com/black-forest-labs/flux-2-pro', '2026-08-18', 'Safety needs separate policy; model-page pricing basis is version-specific.'),
        ('Image', 'fal', 'FLUX.2 Pro $0.03 first output MP + $0.015 additional input/output MP', 'https://fal.ai/models/fal-ai/flux-2-pro', '2026-08-18', 'Default safety checker is NSFW-oriented, not complete children’s policy.'),
        ('Local image', 'ComfyUI', 'Core local runtime; no per-image software price stated', 'https://github.com/Comfy-Org/ComfyUI', '2026-08-18', 'Local GPU, model license, filtering, operations and evaluation remain real costs.'),
        ('FX', 'Bank of England', '17 Aug 2026: £1 = $1.3553; model uses $1 = £0.73784', 'https://www.bankofengland.co.uk/boeapps/database/Rates.asp', '2026-08-18', 'BoE calls these non-official rates; planning translation only.'),
    ]
    for rec in records:
        row += 1
        for col, val in enumerate(rec, 3):
            ws.cell(row, col, val)
            ws.cell(row, col).alignment = Alignment(wrap_text=True, vertical='top')
        comment(ws.cell(row, 6), f'Source accessed {rec[4]}.')
    autofit(ws, max_col=8)
    ws.print_area = f'B2:H{row}'


def add_readme_sheet(wb):
    ws = wb.create_sheet('Read Me', 0)
    apply_layout(ws, 'SynapticGM Own AI vs Paid APIs — Calculator Read Me', 'Fresh research project dated 18 August 2026. This workbook is an incremental operating-cash model, not a pricing commitment or revenue forecast.', '(£ presentation; editable inputs on Assumptions)')
    rows = [
        ('Purpose', 'Compare the incremental 12-month cash cost of five deployment architectures at 100, 1,000 and 10,000 Free MAU. The central decision is whether a small self-hosted Continuity Warden can earn its own operating complexity before anyone considers a self-hosted narrator.'),
        ('Scope', 'Live SynapticGM only. The ledger is truth. The Warden is a narrow classifier, not a custom full narrator. No WOF, no hybrid-climate analysis and no patent analysis are included.'),
        ('How to use', 'Change blue cells in Assumptions. Formula cells are black; cross-sheet links are green. Read Cashflow for base-case totals, Sensitivity for cache/retry/art stress, Image Ladder for image route policy, and Sources for dated vendor evidence.'),
        ('Base workload', '40 turns per Free MAU per month, 5,000 input tokens and 500 output tokens per turn, 30% cache hit, 5% cache-write share when cache is used, 5% API retry rate, and 10× average-to-peak capacity factor.'),
        ('Cost definition', 'Includes narrator inference, gate inference, warm GPU capacity, estimated operator hours, evaluation budget, defined API/hosted fallbacks, and baseline art. Excludes revenue, taxes/VAT, payment processing, existing ledger/platform cost, corporate overhead, storage, egress, capital expenditure and any price discount not published.'),
        ('Interpretation', 'The self-hosted Warden case deliberately uses a warm $0.27/h 24GB Pod, which makes idle capacity visible. Serverless active-only costs are shown separately but do not receive a free low-latency guarantee. The 7B narrator row is explicitly non-equivalent; the 70B row sizes capacity for a 10× peak and 50 output tok/s/GPU.'),
        ('Source policy', 'All price values are public list-price observations from the cited pages, accessed 18 Aug 2026. Some sources are dynamic, model/version-specific or marketplace-based. Recheck before a spend commitment.'),
    ]
    row = 8
    for label, text in rows:
        ws.cell(row, 3, label).font = Font(bold=True)
        ws.cell(row, 4, text)
        ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=10)
        ws.cell(row, 4).alignment = Alignment(wrap_text=True, vertical='top')
        row += 2
    ws.cell(row, 3, 'Workbook check').fill = PatternFill('solid', fgColor='CFE9E0')
    ws.cell(row, 3).font = Font(bold=True)
    row += 1
    ws.cell(row, 3, 'Input color audit')
    ws.cell(row, 4, 'Blue = hardcoded/editable input, Black = calculation on same sheet, Green = cross-sheet reference. Every blue hardcode has a source/assumption comment.')
    ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=10)
    ws.cell(row, 4).alignment = Alignment(wrap_text=True)
    autofit(ws, max_col=10)
    ws.print_area = f'B2:J{row}'


def create_workbook():
    wb = Workbook()
    wb.remove(wb.active)
    add_readme_sheet(wb)
    cells = write_assumptions(wb)
    add_cashflow_sheet(wb, cells)
    add_sensitivity_sheet(wb, cells)
    add_image_sheet(wb, cells)
    add_sources_sheet(wb)
    # Apply common borders/alignment and print-friendly row heights.
    thin_gray = Side(style='thin', color='D9E1F2')
    for ws in wb.worksheets:
        for row in ws.iter_rows(min_row=8, max_row=ws.max_row, min_col=3, max_col=ws.max_column):
            for cell in row:
                if cell.value is not None:
                    cell.border = Border(bottom=thin_gray)
                    if cell.row >= 8 and cell.column >= 3:
                        cell.alignment = Alignment(vertical='top', wrap_text=cell.alignment.wrap_text)
        for r in range(1, ws.max_row + 1):
            ws.row_dimensions[r].height = 18
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = 'auto'
    wb.save(XLSX_PATH)


def export_results(results):
    fields = ['scenario_key', 'mau', 'turns_month', 'narrator_gbp', 'gate_gbp', 'gpu_gbp', 'gpu_count', 'gpu_active_hours', 'gpu_paid_hours', 'gpu_idle_hours', 'operator_gbp', 'eval_gbp', 'fallback_gbp', 'art_gbp', 'monthly_total_gbp', 'annual_total_gbp', 'gbp_per_mau_month', 'notes']
    with CSV_PATH.open('w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=fields)
        writer.writeheader()
        writer.writerows(results)
    with JSON_PATH.open('w', encoding='utf-8') as f:
        json.dump({'assumptions': A, 'rates': RATES, 'results': results}, f, indent=2)


def make_chart(results):
    # Two panels: 1k and 10k MAU. The scale is log because 70B capacity dwarfs the low-cost routes.
    fig, axes = plt.subplots(1, 2, figsize=(16, 7), constrained_layout=True)
    labels = [label.replace(' — ', '\n') for _, label, _ in SCENARIOS]
    colors = ['#5B8FF9', '#61DDAA', '#65789B', '#F6BD16', '#E8684A', '#8E6CFF']
    for ax, mau in zip(axes, [1000, 10000]):
        subset = [r for r in results if r['mau'] == mau]
        vals = [r['annual_total_gbp'] for r in subset]
        bars = ax.barh(labels, vals, color=colors)
        ax.set_title(f'12-Month Incremental Cost — {mau:,} Free MAU')
        ax.set_xlabel('GBP / 12 months (log scale)')
        ax.set_xscale('log')
        ax.grid(axis='x', alpha=0.25)
        for bar, val in zip(bars, vals):
            ax.text(val * 1.05, bar.get_y() + bar.get_height()/2, f'£{val:,.0f}', va='center', fontsize=9)
    fig.suptitle('SynapticGM — Paid API Routing vs Self-Hosted Warden / Full Narrator', fontsize=16, fontweight='bold')
    fig.savefig(PNG_PATH, dpi=200, bbox_inches='tight')
    plt.close(fig)


if __name__ == '__main__':
    results = build_results()
    create_workbook()
    export_results(results)
    make_chart(results)
    print(json.dumps({
        'workbook': str(XLSX_PATH),
        'results_csv': str(CSV_PATH),
        'results_json': str(JSON_PATH),
        'chart': str(PNG_PATH),
        'rows': len(results),
    }, indent=2))
